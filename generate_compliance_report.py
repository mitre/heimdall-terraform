import boto3
from collections import defaultdict 
import json
import os
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from pprint import pprint
import re
import yaml


class ControlInfo():
    def __init__(self, name, compliance):
        """
        Positional Arguments:
            name -- The rule name from AWS
            compliance -- The compliance staus from AWS [COMPLIANT, NON_COMPLIANT, NOT_APPLICABLE, INSUFFICIENT_DATA]
        """
        self.name = name
        self.compliance = compliance
        self._json_dict = None
        self._tags = None
        self._controls = None
        self._test_type = None
        self._cloud_resource = None
        self._category = None
        self._responsibility = None
        self._validation_steps = None
        self._northcom_validated = None

    def __ensure_json_dict(self):
        if not self._json_dict:
            if os.path.isfile(f'rules/{self.name}/parameters.json'):
                with open(f'rules/{self.name}/parameters.json') as file:
                    self._json_dict = json.loads(file.read())
            else:
                self._json_dict = {}

    def __ensure_tags(self):
        self.__ensure_json_dict()
        if not self._tags:
            if 'Tags' in self._json_dict:
                self._tags = json.loads(self._json_dict['Tags'])
            else: 
                self._tags = []

    def __find_tag(self, tag_name):
        """
        Finds a specific tag from the tags array and returns the value

        Positional arguments:
            tag_name -- The tag name to find

        Returns: The tag value or "-"
        """
        self.__ensure_tags()
        for tag_dict in self._tags:
            if tag_dict['Key'] == tag_name:
                return tag_dict['Value']
        return "-"

    @property
    def description(self):
        self.__ensure_json_dict()
        if 'Parameters' in self._json_dict and 'Description' in self._json_dict['Parameters']:
            return self._json_dict['Parameters']['Description']

    @property
    def controls(self):
        if self._controls:
            return self._controls
        if os.path.isfile('rule_mapping.yaml'):
            with open('rule_mapping.yaml') as file:
                yml = yaml.load(file.read(), Loader=yaml.FullLoader)
                if self.name in yml:
                    self._controls = yml[self.name]
                    self._controls.sort()
        if not self._controls:
            self._controls = []
        return self._controls    


    @property
    def test_type(self):
        return self.__find_tag('TestType')

    @property
    def cloud_resource(self):
        return self.__find_tag('CloudResource')

    @property
    def category(self):
        return self.__find_tag('Category')

    @property
    def responsibility(self):
        return self.__find_tag('Responsibility')

    @property
    def validation_steps(self):
        return self.__find_tag('ValidationSteps')

    @property
    def northcom_validated(self):
        return self.__find_tag('USNORTHCOMValidated')
    
    @property
    def is_tracked_rule(self):
        return self.controls and self.description
    

    def get_rule_compliance_row(self):
        return [
            self.name,
            self.description,
            ',\n'.join(self.controls),
            self.compliance,
            self.test_type,
            self.cloud_resource,
            self.category,
            self.responsibility,
            self.validation_steps,
            self.northcom_validated,
        ]

    def __repr__(self):
        return f'ControlInfo({self.name}, {self.compliance})'


def get_all_config_rule_compliance():
    """
    Fetch the compliance status of all the config rules for the AWS account
    https://boto3.amazonaws.com/v1/documentation/api/latest/reference/services/config.html#ConfigService.Client.get_compliance_details_by_config_rule
    
    Returns: Config rule status in the format { 'Rule_Name': 'COMPLIANT', ... }
    """
    def process_rules(response):
        if 'ComplianceByConfigRules' in response:
            return [process_rule(rule_status) for rule_status in response['ComplianceByConfigRules']]
        return []

    def process_rule(rule_status):
        rule_name = rule_status['ConfigRuleName']
        if 'Compliance' in rule_status and 'ComplianceType' in rule_status['Compliance']:
            return ControlInfo(rule_name, rule_status['Compliance']['ComplianceType'])

    client = boto3.client('config')
    response = client.describe_compliance_by_config_rule()
    all_control_info = process_rules(response)

    while 'NextToken' in response and response['NextToken']:
        response = client.describe_compliance_by_config_rule(NextToken=response['NextToken'])
        all_control_info.extend(process_rules(response))

    return all_control_info


def get_control_link(control_name):
    match = re.match('^[A-Z]{2}-[0-9]{2}', control_name)
    if match:
        match = match.group(0)
        return f'https://nvd.nist.gov/800-53/Rev4/control/{match[0:2]}-{str(int(match[3:5]))}'
    return ''


def get_control_compliance_rows(all_control_info):
    def default():
        return { 
            'Control Link': '', 
            'Compliant Rules': [], 
            'Non-Compliant Rules': [], 
            'Insufficient Data Rules': [],
            'Not Applicable Rules': []
        }
    compliance_dict = defaultdict(default)
    for control_info in all_control_info:
        if not control_info.is_tracked_rule:
            continue

        for control in control_info.controls:
            target_dict = compliance_dict[control]

            # Set compliance
            if control_info.compliance == 'COMPLIANT':
                target_dict['Compliant Rules'].append(control_info.name)
            if control_info.compliance == 'NON_COMPLIANT':
                target_dict['Non-Compliant Rules'].append(control_info.name)
            if control_info.compliance == 'INSUFFICIENT_DATA':
                target_dict['Insufficient Data Rules'].append(control_info.name)
            if control_info.compliance == 'NOT_APPLICABLE':
                target_dict['Not Applicable Rules'].append(control_info.name)

            # Set control link 
            target_dict['Control Link'] = get_control_link(control)

    rows = [
        [
            key,
            value['Control Link'],
            ',\n'.join(value['Compliant Rules']),
            ',\n'.join(value['Non-Compliant Rules']),
            ',\n'.join(value['Insufficient Data Rules']),
            ',\n'.join(value['Not Applicable Rules'])
        ] for key, value in compliance_dict.items()
    ]

    rows.sort(key=lambda x: x[0])

    return rows


def export_compliance(all_control_info):
    wb = load_workbook(filename='Compliance-Results-TEMPLATE.xlsx')
    rule_compliance_sheet = wb['Rule Compliance']
    for row in [ci.get_rule_compliance_row() for ci in all_control_info if ci.is_tracked_rule]: 
        rule_compliance_sheet.append(row)
    control_compliance_sheet = wb['Control Compliance']
    for row in get_control_compliance_rows(all_control_info):
        control_compliance_sheet.append(row)

    alignment = Alignment(
        horizontal='center',
        vertical='center',
        wrap_text=True
    )
    apply_alignment_to_all_sheet_cells(rule_compliance_sheet, alignment)
    apply_alignment_to_all_sheet_cells(control_compliance_sheet, alignment)

    wb.save(filename='Compliance-Results.xlsx')

def apply_alignment_to_all_sheet_cells(sheet, alignment):
    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = alignment

all_control_info = get_all_config_rule_compliance()
export_compliance(all_control_info)
# import pdb; pdb.set_trace()

